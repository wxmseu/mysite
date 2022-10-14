from django.shortcuts import render, redirect, HttpResponse
from myblog import models
from openpyxl import load_workbook


def depart_list(request):
    """部门列表"""
    depart_list_info = models.Department.objects.all()
    return render(request, 'myblog/depart_list.html', {'depart_list_info': depart_list_info})


def depart_add(request):
    """添加部门"""
    if request.method == 'GET':
        return render(request, 'myblog/depart_add.html')
    department = request.POST.get('department')
    models.Department.objects.create(title=department)
    return redirect('/depart/list/')


def depart_delete(request):
    """删除部门"""
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect('/depart/list/')


def depart_edit(request, nid):
    """编辑部门"""
    if request.method == 'GET':
        edit_depart = models.Department.objects.filter(id=nid).first()
        return render(request, 'myblog/depart_edit.html', {'edit_depart': edit_depart})

    new_depart = request.POST.get('new_depart')
    models.Department.objects.filter(id=nid).update(title=new_depart)
    return redirect('/depart/list/')


def depart_upload(request):
    # 1.获取用户上传的文件
    file_obj = request.FILES.get('exc')
    # 2. 对象传递给openpyxl,由openpyxl读取文件内容
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    # 3.循环读取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exist = models.Department.objects.filter(title=text).exists()
        if not exist:
            models.Department.objects.create(title=text)
    return redirect('/depart/list/')
